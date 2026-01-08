# app.py
# IMPORTANT:
# - Do NOT include any markdown backticks (```python) in this file.
# - For Streamlit Cloud: add openpyxl in requirements.txt to enable Excel export (.xlsx).

import re
import time
from io import BytesIO
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
import streamlit as st

# Optional Excel export (app still runs without it)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False


# =============================================================================
# Constants / Regex
# =============================================================================

DEFAULT_HEADERS = {
    "User-Agent": "Inovestor-Directory-Extractor/0.5.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

TEAM_PAGE_TEXT_PAT = re.compile(r"\b(our team|notre équipe|team members|membres de l[' ]équipe)\b", re.I)
CONTACT_PAGE_TEXT_PAT = re.compile(r"\b(contact|contactez-nous|nous joindre|communiqu|communicat)\b", re.I)

EMAIL_RE = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.I)
PHONE_RE = re.compile(r"(\+?\d[\d\-\s().]{7,}\d)")
POSTAL_CA_RE = re.compile(r"\b[ABCEGHJ-NPRSTVXY]\d[ABCEGHJ-NPRSTV-Z][ -]?\d[ABCEGHJ-NPRSTV-Z]\d\b", re.I)

BANNED_WORDS = set("""
contact communiquer communique contactez nous joindre
approach commitment services service produits product planning planification patrimoine
privabanque bio biographie team accueil home
wealth investment community partners partner
successoraux fiduciaires fondée founded savoir plus visitez visit
email call connect discovery process additional specialist specialists
""".split())

PARTICLES = set([
    "de", "du", "des", "la", "le", "da", "di", "del", "della", "van", "von", "der", "den",
    "st", "ste", "saint", "sainte", "mc", "mac", "o'"
])

ROLE_WORDS = {
    "senior", "branch", "administrator", "admin", "assistant", "associate", "advisor", "adviser",
    "manager", "director", "president", "vp", "vice", "consultant", "specialist", "partner",
    "investment", "portfolio", "financial", "wealth", "planner", "planning",
    "conseiller", "conseillère", "placement", "gestionnaire", "directeur", "président", "adjointe", "adjoint",
    "client", "service", "representative", "représentant", "représentante"
}

JUNK_PHRASES = {
    "our branch team", "notre équipe de succursale", "our team", "notre équipe",
    "email us", "call us", "contact us", "let's connect", "lets connect",
    "additional td specialists", "a unique discovery process", "discovery process"
}

TD_STOP_MARKERS = {"Additional TD Specialists", "Spécialistes TD additionnels", "Additional TD specialists"}
TD_SOCIAL_MARKERS = {"social links", "liens sociaux"}


# =============================================================================
# URL helpers
# =============================================================================

def norm_url(u: str) -> str:
    p = urlparse(u)
    return p._replace(fragment="", query="").geturl()

def same_domain(a: str, b: str) -> bool:
    return urlparse(a).netloc.lower() == urlparse(b).netloc.lower()

def extract_links(html: str, base_url: str):
    soup = BeautifulSoup(html, "html.parser")
    out = []
    for a in soup.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        if not href:
            continue
        abs_url = norm_url(urljoin(base_url, href))
        text = a.get_text(" ", strip=True)
        out.append((text, abs_url))
    return out

def page_title(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    h1 = soup.find("h1")
    if h1 and h1.get_text(strip=True):
        return h1.get_text(" ", strip=True)
    if soup.title and soup.title.string:
        return soup.title.string.strip()
    return ""

def find_best_link(links, base_url: str, pattern: re.Pattern):
    candidates = []
    for text, url in links:
        if not same_domain(url, base_url):
            continue
        if pattern.search(text or "") or pattern.search(url or ""):
            candidates.append((text, url))
    candidates.sort(key=lambda x: len(urlparse(x[1]).path))
    return candidates[0][1] if candidates else ""


# =============================================================================
# Domain detectors
# =============================================================================

def is_td_url(u: str) -> bool:
    return "advisors.td.com" in (urlparse(u).netloc or "").lower()

def is_desjardins_url(u: str) -> bool:
    return "desjardins.com" in (urlparse(u).netloc or "").lower()

def is_cibc_wg_url(u: str) -> bool:
    return "woodgundyadvisors.cibc.com" in (urlparse(u).netloc or "").lower()


# =============================================================================
# Requests session + polite_get (accent-safe decoding + cache)
# =============================================================================

SESSION = requests.Session()
SESSION.headers.update(DEFAULT_HEADERS)
_PAGE_CACHE = {}

def polite_get(url: str, sleep_s: float = 0.75, timeout: int = 25, retries: int = 3):
    """Polite GET with retry/backoff + safer decoding (accents) + cache."""
    ukey = norm_url(url)
    if ukey in _PAGE_CACHE:
        return _PAGE_CACHE[ukey]

    time.sleep(max(0.0, sleep_s))
    last_err = None
    for attempt in range(retries):
        try:
            r = SESSION.get(url, timeout=timeout, allow_redirects=True)
            r.raise_for_status()

            # Fix accents: requests often defaults to ISO-8859-1
            enc = (r.encoding or "").lower()
            if not enc or enc == "iso-8859-1":
                r.encoding = r.apparent_encoding or "utf-8"

            html = r.text
            final_url = r.url
            _PAGE_CACHE[ukey] = (html, final_url)

            if len(_PAGE_CACHE) > 900:
                _PAGE_CACHE.pop(next(iter(_PAGE_CACHE)))

            return html, final_url
        except Exception as e:
            last_err = e
            time.sleep(1.25 * (attempt + 1))
    raise last_err


# =============================================================================
# Name / Role helpers
# =============================================================================

def clean_person_name(raw: str) -> str:
    s = str(raw or "")
    s = s.replace("\u00A0", " ").replace("’", "'")
    s = re.sub(r"\([^)]*\)", "", s).strip()
    if "," in s:
        s = s.split(",", 1)[0].strip()
    s = re.sub(r"[^A-Za-zÀ-ÿ\-\s'\.]", "", s)
    s = re.sub(r"\s{2,}", " ", s).strip(" -–—|")
    return s

def is_valid_person_name(raw: str) -> bool:
    s = clean_person_name(raw)
    if not s or re.search(r"\d", s):
        return False
    if s.lower().strip() in JUNK_PHRASES:
        return False

    tokens = s.split()
    if len(tokens) < 2 or len(tokens) > 6:
        return False

    low_tokens = [t.lower().strip(".") for t in tokens]
    if any(t in BANNED_WORDS for t in low_tokens):
        return False

    caps = 0
    for t in tokens:
        tl = t.lower().strip(".")
        if tl in PARTICLES:
            continue
        if re.match(r"^[A-ZÀ-Ý]", t):
            caps += 1
        else:
            return False
    return caps >= 2

def canon_name(raw: str) -> str:
    return re.sub(r"[^a-z]+", "", clean_person_name(raw).lower())

def _canon(s: str) -> str:
    return re.sub(r"[^a-z]+", "", (s or "").lower())

def is_likely_role(text: str, person_name: str = "") -> bool:
    if not text:
        return False
    t = re.sub(r"\s+", " ", text).strip(" -|•·")
    if len(t) < 2 or len(t) > 120:
        return False
    tl = t.lower()
    if tl in JUNK_PHRASES:
        return False
    if EMAIL_RE.search(t) or PHONE_RE.search(t):
        return False
    if person_name and _canon(t) == _canon(person_name):
        return False

    toks = re.findall(r"[A-Za-zÀ-ÿ']+", tl)
    return any(tok in ROLE_WORDS for tok in toks)

def _first_email(email_field: str) -> str:
    s = (email_field or "").strip()
    if not s:
        return ""
    return s.split(";")[0].strip().lower()

def _digits_phone(phone_field: str) -> str:
    s = (phone_field or "").strip()
    if not s:
        return ""
    return re.sub(r"\D+", "", s)[:15]

def _normalize_phone_list(phone_candidates):
    by_digits = {}
    for p in phone_candidates:
        s = re.sub(r"\s+", " ", (p or "")).strip()
        digs = re.sub(r"\D+", "", s)
        if len(digs) < 10:
            continue
        score = 0
        if "(" in s and ")" in s:
            score += 3
        if " " in s or "-" in s:
            score += 1
        if len(s) <= 18:
            score += 1
        if digs not in by_digits or score > by_digits[digs][0]:
            by_digits[digs] = (score, s)
    out = [v[1] for v in sorted(by_digits.values(), key=lambda x: -x[0])]
    return out[:3]


# =============================================================================
# TD logic (Directory -> Advisors + Teams)
# =============================================================================

def td_root_from_any_td_url(u: str) -> str:
    p = urlparse(u)
    parts = [x for x in p.path.split("/") if x]
    if not parts:
        return f"{p.scheme}://{p.netloc}/"
    slug = parts[0]
    return f"{p.scheme}://{p.netloc}/{slug}/"

def _td_is_one_segment_root(u: str) -> bool:
    p = urlparse(u)
    parts = [x for x in p.path.strip("/").split("/") if x]
    return len(parts) == 1

def _norm_heading_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").replace("\u00A0", " ")).strip().lower()

def td_is_directory_page(soup: BeautifulSoup) -> bool:
    headings = {_norm_heading_text(h.get_text(" ", strip=True)) for h in soup.find_all(["h2", "h3", "h4"])}
    has_advisors = any(x in headings for x in {"advisors", "advisor", "conseillers", "conseiller"})
    has_teams = any(x in headings for x in {"teams", "team", "équipes", "equipes", "équipe", "equipe"})
    return has_advisors and has_teams

def td_extract_links_under_heading(soup: BeautifulSoup, base_url: str, heading_set: set):
    """Collect links under the heading block. Works for typical 'Advisors' / 'Teams' sections."""
    for h in soup.find_all(["h2", "h3", "h4"]):
        ht = _norm_heading_text(h.get_text(" ", strip=True))
        if ht not in heading_set:
            continue

        links = []
        for sib in h.find_next_siblings():
            if getattr(sib, "name", None) in ["h2", "h3", "h4"]:
                break
            for a in sib.find_all("a", href=True):
                t = a.get_text(" ", strip=True)
                u = norm_url(urljoin(base_url, a.get("href")))
                links.append((t, u))

        # fallback: tabbed layouts sometimes break siblings, so scan further down
        if not links:
            for a in h.find_all_next("a", href=True, limit=450):
                t = a.get_text(" ", strip=True)
                u = norm_url(urljoin(base_url, a.get("href")))
                links.append((t, u))

        out, seen = [], set()
        for t, u in links:
            if not is_td_url(u):
                continue
            root = td_root_from_any_td_url(u)
            if not _td_is_one_segment_root(root):
                continue
            k = root.lower()
            if k in seen:
                continue
            seen.add(k)
            out.append((t, root))
        return out
    return []

def td_scan_all_one_segment_roots(html: str, base_url: str):
    """Hard fallback: find all TD one-segment roots on the directory page."""
    links = extract_links(html, base_url)
    roots, seen = [], set()

    branch_slug = (urlparse(base_url).path.strip("/").split("/")[0].lower()
                   if urlparse(base_url).path.strip("/") else "")

    for text, u in links:
        if not is_td_url(u):
            continue
        root = td_root_from_any_td_url(u)
        if not _td_is_one_segment_root(root):
            continue
        seg = urlparse(root).path.strip("/").lower()
        if not seg or seg == branch_slug:
            continue
        k = root.lower()
        if k in seen:
            continue
        seen.add(k)
        roots.append((text or root, root))
    return roots

def td_detect_single_root_kind(html: str) -> str:
    """Advisor profiles usually contain 'Part of / Fait partie de' link."""
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=True):
        t = a.get_text(" ", strip=True)
        if re.match(r"^(Part of|Fait partie de)\b", t, re.I):
            return "advisor"
    return "team"

def td_extract_part_of_team(html: str, base_url: str):
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=True):
        t = a.get_text(" ", strip=True)
        if re.match(r"^(Part of|Fait partie de)\b", t, re.I):
            team_name = re.sub(r"^(Part of|Fait partie de)\s*", "", t, flags=re.I).strip()
            team_url = norm_url(urljoin(base_url, a.get("href")))
            if is_td_url(team_url):
                return team_name, td_root_from_any_td_url(team_url)
    return "", ""

def td_guess_meet_the_team_url(root_final: str, sleep_s: float):
    guesses = [
        "meet-the-team.htm", "meet-the-team.html", "meet-the-team",
        "meet-the-advisors.htm", "meet-the-advisors.html"
    ]
    base = root_final.rstrip("/") + "/"
    for g in guesses:
        u = urljoin(base, g)
        try:
            _, final_u = polite_get(u, sleep_s=sleep_s)
            return final_u
        except Exception:
            continue
    return ""

def td_extract_person_from_profile(html: str, base_url: str):
    soup = BeautifulSoup(html, "html.parser")
    h1 = soup.find("h1")
    if not h1:
        return None

    name = clean_person_name(h1.get_text(" ", strip=True))
    if not is_valid_person_name(name):
        return None

    # Role (best-effort from nearby lines)
    lines = [x.strip() for x in soup.get_text("\n", strip=True).split("\n") if x.strip()]
    role_lines = []
    try:
        idx = next(i for i, x in enumerate(lines) if clean_person_name(x) == name)
    except StopIteration:
        idx = 0

    stop_words = {"phone", "toll free", "fax", "mobile", "office location", "contact",
                  "courriel", "téléphone", "telephone", "adresse"}
    for line in lines[idx + 1: idx + 24]:
        if line.lower() in stop_words:
            break
        if is_likely_role(line, name):
            role_lines.append(line)
    role = " / ".join(dict.fromkeys(role_lines))[:120]

    # Email
    emails = []
    for a in soup.select('a[href^="mailto:"]'):
        href = a.get("href", "")
        e = href.split("mailto:", 1)[-1].split("?", 1)[0].strip()
        if e and e not in emails:
            emails.append(e)
    if not emails:
        for m in EMAIL_RE.findall(soup.get_text(" ", strip=True)):
            if m not in emails:
                emails.append(m)

    # Phones
    phones = []
    for a in soup.select('a[href^="tel:"]'):
        href = a.get("href", "")
        p = href.split("tel:", 1)[-1].strip()
        if p and p not in phones:
            phones.append(p)
    if not phones:
        for m in PHONE_RE.findall(soup.get_text(" ", strip=True)):
            phones.append(m)
    phones = _normalize_phone_list(phones)

    # Address
    address = ""
    text_lines = [x.strip() for x in soup.get_text("\n", strip=True).split("\n") if x.strip()]
    for i, line in enumerate(text_lines):
        if line.lower() in {"office location", "adresse du bureau"}:
            if i + 1 < len(text_lines):
                address = text_lines[i + 1]
            break
    if not address:
        for i, line in enumerate(text_lines):
            if POSTAL_CA_RE.search(line):
                start = max(0, i - 2)
                end = min(len(text_lines), i + 2)
                address = " | ".join(text_lines[start:end])
                break

    return {
        "advisor_name": name,
        "advisor_role": role,
        "advisor_email": "; ".join(emails[:3]),
        "advisor_phone": "; ".join(phones[:3]),
        "advisor_address": address,
        "advisor_profile_url": norm_url(base_url),
        "source": "td_profile",
    }

def td_extract_people_from_meet_page(html: str):
    """
    TD team roster pages are pretty unstructured; this parses "stripped_strings"
    and groups blocks based on when a new valid name appears after contact info.
    """
    soup = BeautifulSoup(html, "html.parser")
    strings = [s.strip().replace("\u00A0", " ") for s in soup.stripped_strings if s and s.strip()]

    trimmed = []
    for s in strings:
        if s in TD_STOP_MARKERS:
            break
        trimmed.append(s)

    entries, cur = [], []

    def has_any_contact(buf):
        txt = " ".join(buf)
        return bool(EMAIL_RE.search(txt) or PHONE_RE.search(txt))

    def looks_like_person_line(x: str) -> bool:
        nm = clean_person_name(x)
        return is_valid_person_name(nm) and (nm.lower() not in JUNK_PHRASES)

    for s in trimmed:
        if s.strip().lower() in TD_SOCIAL_MARKERS:
            if cur:
                entries.append(cur)
            cur = []
            continue
        if s.lower() == "photo":
            continue

        # new person boundary
        if cur and has_any_contact(cur) and looks_like_person_line(s):
            entries.append(cur)
            cur = [s]
            continue

        cur.append(s)

    if cur:
        entries.append(cur)

    people = []
    for buf in entries:
        # name
        name = ""
        for x in buf:
            nm = clean_person_name(x)
            if is_valid_person_name(nm):
                name = nm
                break
        if not name:
            continue

        # emails
        emails = []
        for x in buf:
            for m in EMAIL_RE.findall(x):
                if m not in emails:
                    emails.append(m)

        # phones
        phone_candidates = []
        for x in buf:
            for m in PHONE_RE.findall(x):
                phone_candidates.append(m)
        phones = _normalize_phone_list(phone_candidates)

        # role (lines after name until contact)
        role_lines, hit_name = [], False
        for x in buf:
            if clean_person_name(x) == name:
                hit_name = True
                continue
            if not hit_name:
                continue
            if EMAIL_RE.search(x) or PHONE_RE.search(x):
                break
            if is_likely_role(x, name):
                role_lines.append(x)

        role = " / ".join(dict.fromkeys(role_lines))[:120]

        people.append({
            "advisor_name": name,
            "advisor_role": role,
            "advisor_email": "; ".join(emails[:3]),
            "advisor_phone": "; ".join(phones[:3]),
            "advisor_address": "",
            "advisor_profile_url": "",
            "source": "td_meet_the_team",
        })

    # de-dupe within roster page
    seen, out = set(), []
    for p in people:
        key = (_first_email(p.get("advisor_email")) or canon_name(p.get("advisor_name") or ""))
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
    return out

def td_fetch_people(url: str, sleep_s: float):
    html, final_url = polite_get(url, sleep_s=sleep_s)
    path = (urlparse(final_url).path or "").lower()

    if "meet-the-team" in path or "meet-the-advisors" in path:
        return td_extract_people_from_meet_page(html), final_url

    p = td_extract_person_from_profile(html, final_url)
    if p:
        return [p], final_url

    # last chance: roster parser if page text suggests it
    low = html.lower()
    if ("meet the team" in low) or ("rencontrez l" in low) or ("rencontrez l’équipe" in low) or ("rencontrez l'equipe" in low):
        roster = td_extract_people_from_meet_page(html)
        if roster:
            return roster, final_url

    return [], final_url

def discover_td_targets(seed_url: str, sleep_s: float):
    """
    TD directory pages have 2 sections: Advisors + Teams.
    We collect both lists, then later:
      - crawl each advisor profile
      - crawl each team roster
    """
    html, final_url = polite_get(seed_url, sleep_s=sleep_s)
    soup = BeautifulSoup(html, "html.parser")

    if td_is_directory_page(soup):
        advisors = td_extract_links_under_heading(
            soup, final_url, {"advisors", "advisor", "conseillers", "conseiller"}
        )
        teams = td_extract_links_under_heading(
            soup, final_url, {"teams", "team", "équipes", "equipes", "équipe", "equipe"}
        )

        rows = []
        for t, u in advisors:
            rows.append({"branch_seed_url": seed_url, "target_url": u, "link_text": t, "td_kind": "advisor"})
        for t, u in teams:
            rows.append({"branch_seed_url": seed_url, "target_url": u, "link_text": t, "td_kind": "team"})

        df = pd.DataFrame(rows)
        if not df.empty:
            return df.drop_duplicates(subset=["target_url", "td_kind"]).reset_index(drop=True)

    # fallback: grab all TD roots and infer kind later
    roots = td_scan_all_one_segment_roots(html, final_url)
    df = pd.DataFrame([{"branch_seed_url": seed_url, "target_url": u, "link_text": t, "td_kind": ""} for t, u in roots])
    if df.empty:
        root = td_root_from_any_td_url(final_url)
        df = pd.DataFrame([{"branch_seed_url": seed_url, "target_url": root, "link_text": "seed", "td_kind": ""}])
    return df.drop_duplicates(subset=["target_url"]).reset_index(drop=True)


# =============================================================================
# Desjardins discovery
# =============================================================================

DESJARDINS_TEAM_LINK_RE = re.compile(r"/find-us/desjardins-securities-team/[^/?#]+\.html$", re.I)

def discover_desjardins_team_pages(seed_url: str, sleep_s: float):
    html, final_url = polite_get(seed_url, sleep_s=sleep_s)
    links = extract_links(html, final_url)

    candidates = []
    if DESJARDINS_TEAM_LINK_RE.search(urlparse(final_url).path):
        candidates.append({"branch_seed_url": seed_url, "target_url": norm_url(final_url), "link_text": "seed", "td_kind": ""})

    for text, u in links:
        if not is_desjardins_url(u):
            continue
        if not DESJARDINS_TEAM_LINK_RE.search(urlparse(u).path):
            continue
        t = (text or "").strip()
        if t.lower().startswith("view profile") or t.lower().startswith("voir le profil"):
            continue
        candidates.append({"branch_seed_url": seed_url, "target_url": norm_url(u), "link_text": t or u, "td_kind": ""})

    df = pd.DataFrame(candidates)
    if df.empty:
        return pd.DataFrame(columns=["branch_seed_url", "target_url", "link_text", "td_kind"])
    return df.drop_duplicates(subset=["target_url"]).reset_index(drop=True)


# =============================================================================
# CIBC Wood Gundy discovery support + common discover
# =============================================================================

def branch_slug_from_url(url: str) -> str:
    parts = urlparse(url).path.strip("/").split("/")
    if len(parts) >= 2 and parts[0].lower() == "web":
        return parts[1].lower()
    return ""

def is_true_team_root(url: str, branch_slug: str) -> bool:
    path = urlparse(url).path.strip("/")
    if not path:
        return False
    parts = path.split("/")
    if len(parts) != 1:
        return False
    seg = parts[0].lower()
    if branch_slug and seg == branch_slug:
        return False
    if seg in {"web", "home", "contact", "services", "produits", "products"}:
        return False
    return True

def discover_targets_from_seed(seed_url: str, sleep_s: float):
    if is_td_url(seed_url):
        return discover_td_targets(seed_url, sleep_s=sleep_s)

    if is_desjardins_url(seed_url):
        return discover_desjardins_team_pages(seed_url, sleep_s=sleep_s)

    # CIBC WG: discover team roots from the teams hub page
    html, final_url = polite_get(seed_url, sleep_s=sleep_s)
    links = extract_links(html, final_url)

    if "our-investment-advisors-and-their-teams" not in final_url.lower():
        for _, u in links:
            if "our-investment-advisors-and-their-teams" in (u or "").lower():
                html, final_url = polite_get(u, sleep_s=sleep_s)
                links = extract_links(html, final_url)
                break

    branch_slug = branch_slug_from_url(final_url)
    candidates = []
    for text, u in links:
        if not same_domain(u, final_url):
            continue
        if is_true_team_root(u, branch_slug):
            candidates.append({"branch_seed_url": seed_url, "target_url": u, "link_text": text, "td_kind": ""})

    df = pd.DataFrame(candidates)
    if df.empty:
        return pd.DataFrame(columns=["branch_seed_url", "target_url", "link_text", "td_kind"])
    return df.drop_duplicates(subset=["target_url"]).reset_index(drop=True)
# =============================================================================
# Slug helper (internal; NOT exported)
# =============================================================================

def to_team_slug(team_root_url: str) -> str:
    p = urlparse(team_root_url)
    host = (p.netloc or "").lower()
    parts = [x for x in p.path.strip("/").split("/") if x]

    # TD
    if "advisors.td.com" in host and parts and parts[0]:
        return parts[0].lower()

    # Desjardins
    if "desjardins.com" in host and parts:
        last = parts[-1].lower().replace(".html", "")
        last = last.replace("_", "-")
        last = re.sub(r"[^a-z0-9\-]+", "-", last)
        last = re.sub(r"-{2,}", "-", last).strip("-")
        return last

    # CIBC (one-segment team root)
    seg = parts[0] if parts else ""
    seg = seg.replace("_", "-")
    seg = re.sub(r"[^A-Za-z0-9\-]+", "-", seg)
    seg = re.sub(r"-{2,}", "-", seg).strip("-")
    return seg.lower()


# =============================================================================
# Generic people extraction (non-TD)
# =============================================================================

def looks_like_name(name: str) -> bool:
    s = (name or "").strip()
    if not s:
        return False
    sl = s.lower()
    if sl in JUNK_PHRASES:
        return False
    parts = s.split()
    if len(parts) < 2 or len(parts) > 6:
        return False
    if EMAIL_RE.search(s) or PHONE_RE.search(s):
        return False
    if not re.match(r"^[A-Za-zÀ-ÿ]", s):
        return False
    return True

def extract_contact_from_block(block: BeautifulSoup):
    emails = []
    for a in block.select('a[href^="mailto:"]'):
        href = a.get("href", "")
        e = href.split("mailto:", 1)[-1].split("?", 1)[0].strip()
        if e and e not in emails:
            emails.append(e)
    if not emails:
        for m in EMAIL_RE.findall(block.get_text(" ", strip=True)):
            if m not in emails:
                emails.append(m)

    phones = []
    for a in block.select('a[href^="tel:"]'):
        href = a.get("href", "")
        p = href.split("tel:", 1)[-1].strip()
        if p and p not in phones:
            phones.append(p)
    if not phones:
        txt = block.get_text(" ", strip=True)
        for m in PHONE_RE.findall(txt):
            m = re.sub(r"\s+", " ", m).strip()
            if m and m not in phones:
                phones.append(m)

    address = ""
    txt_lines = [x.strip() for x in block.get_text("\n", strip=True).split("\n") if x.strip()]
    for i, line in enumerate(txt_lines):
        if POSTAL_CA_RE.search(line):
            start = max(0, i - 2)
            end = min(len(txt_lines), i + 2)
            address = " | ".join(txt_lines[start:end])
            break

    return {
        "advisor_email": "; ".join(emails[:3]),
        "advisor_phone": "; ".join(_normalize_phone_list(phones)[:3]),
        "advisor_address": address
    }

def extract_people_from_page(html: str, base_url: str):
    soup = BeautifulSoup(html, "html.parser")
    people = []

    for h in soup.find_all(["h2", "h3", "h4", "h5"]):
        raw = h.get_text(" ", strip=True)
        name = re.sub(r"\s+", " ", raw or "").strip()
        if not looks_like_name(name):
            continue

        block = h
        for _ in range(4):
            if block.parent is None:
                break
            block = block.parent
            if block.select_one('a[href^="mailto:"]') or block.select_one('a[href^="tel:"]'):
                break

        role = ""
        for sib in h.find_next_siblings(limit=6):
            txt = sib.get_text(" ", strip=True)
            if is_likely_role(txt, name):
                role = txt
                break

        contact = extract_contact_from_block(block)
        profile_url = ""
        a = h.find("a", href=True)
        if a:
            profile_url = norm_url(urljoin(base_url, a.get("href")))

        people.append({
            "advisor_name": clean_person_name(name),
            "advisor_role": role,
            "advisor_profile_url": profile_url,
            **contact,
            "source": "heuristic_block"
        })

    # page-level dedupe
    seen = set()
    out = []
    for p in people:
        k = (_first_email(p.get("advisor_email")) or canon_name(p.get("advisor_name") or ""), _digits_phone(p.get("advisor_phone") or ""))
        if k in seen:
            continue
        seen.add(k)
        out.append(p)
    return out

def fetch_people(url: str, sleep_s: float):
    if is_td_url(url):
        return td_fetch_people(url, sleep_s=sleep_s)

    html, final_url = polite_get(url, sleep_s=sleep_s)
    people = extract_people_from_page(html, final_url)
    return people, final_url


# =============================================================================
# Post-processing / Global de-dupe (ensures same person is NOT on 2 lines)
# - If a person appears in multiple teams, we merge team names into one cell.
# =============================================================================

BASE_COLS = [
    "branch_seed_url",
    "team_root_url", "team_slug", "team_name",
    "team_page_url", "contact_page_url",
    "advisor_name", "advisor_role", "advisor_email", "advisor_phone",
    "advisor_address", "advisor_profile_url",
    "source", "source_page_used"
]

def _ensure_cols(df: pd.DataFrame, cols, fill=""):
    for c in cols:
        if c not in df.columns:
            df[c] = fill
    return df

def post_process_directory(df_out: pd.DataFrame, drop_no_contact=False) -> pd.DataFrame:
    df = df_out.copy()
    df = _ensure_cols(df, BASE_COLS, fill="")

    df["advisor_name"] = df["advisor_name"].apply(clean_person_name)
    df = df[df["advisor_name"].apply(is_valid_person_name)].copy()
    if df.empty:
        return pd.DataFrame(columns=BASE_COLS)

    # rank rows: keep richer info first
    def score_row(r):
        score = 0
        for c in ["advisor_email", "advisor_phone", "advisor_address", "advisor_profile_url"]:
            if str(r.get(c) or "").strip():
                score += 1
        role = str(r.get("advisor_role") or "").strip()
        if is_likely_role(role, str(r.get("advisor_name") or "")):
            score += 1
        return score

    df["_score"] = df.apply(score_row, axis=1)
    df = df.sort_values("_score", ascending=False)

    def person_key(r):
        prof = (r.get("advisor_profile_url") or "").strip().lower()
        if prof:
            return "p:" + prof
        em = _first_email(r.get("advisor_email") or "")
        if em:
            return "e:" + em
        ph = _digits_phone(r.get("advisor_phone") or "")
        if ph:
            return "t:" + ph
        return "n:" + canon_name(r.get("advisor_name") or "")

    df["person_key"] = df.apply(person_key, axis=1)

    merged = []
    for _, g in df.groupby("person_key", sort=False):
        base = g.iloc[0].to_dict()

        # merge all team names into one cell (prevents duplicate lines)
        team_names = [str(x).strip() for x in g["team_name"].tolist() if str(x).strip()]
        base["team_name"] = "; ".join(dict.fromkeys(team_names))

        # keep best info for contact fields
        for col in ["advisor_role", "advisor_email", "advisor_phone", "advisor_address", "advisor_profile_url"]:
            vals = [v for v in g[col].tolist() if pd.notna(v) and str(v).strip()]
            if col == "advisor_role":
                nm = base.get("advisor_name", "")
                vals = [v for v in vals if is_likely_role(str(v), nm)]
            base[col] = vals[0] if vals else (base.get(col) or "")

        merged.append(base)

    out = pd.DataFrame(merged)
    out = _ensure_cols(out, BASE_COLS, fill="")
    out = out.drop(columns=["_score", "person_key"], errors="ignore")

    if drop_no_contact:
        out = out[
            (out["advisor_email"].fillna("").str.strip() != "") |
            (out["advisor_phone"].fillna("").str.strip() != "")
        ].copy()

    return out.reset_index(drop=True)


# =============================================================================
# Excel export (auto-fit + team banding + sheets per input link)
# - Exported sheets DO NOT include branch_seed_url nor team_slug (your request)
# =============================================================================

def _safe_sheet_name(name: str) -> str:
    s = str(name or "").strip()
    s = re.sub(r"[\[\]\:\*\?\/\\]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return (s or "Sheet")[:31]

def _autofit_columns(ws, max_width: int = 52, min_width: int = 10):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))

def _apply_team_banding(ws, team_col_idx: int, start_row: int, end_row: int):
    fills = [
        PatternFill("solid", fgColor="F5F7FF"),
        PatternFill("solid", fgColor="F7F7F7"),
        PatternFill("solid", fgColor="F4FFF7"),
        PatternFill("solid", fgColor="FFF7F4"),
    ]
    team_to_fill, fill_idx = {}, 0

    for r in range(start_row, end_row + 1):
        team_val = ws.cell(row=r, column=team_col_idx).value
        team_raw = str(team_val or "").strip()
        team_key = team_raw.split(";")[0].strip().lower() if team_raw else ""

        if team_key not in team_to_fill:
            team_to_fill[team_key] = fills[fill_idx % len(fills)]
            fill_idx += 1

        fill = team_to_fill[team_key]
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

def build_styled_xlsx(df_full: pd.DataFrame, sheet_group_col: str, out_cols: list, band_by_col: str) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    sheets = [("All", df_full)]
    if sheet_group_col in df_full.columns:
        for seed, g in df_full.groupby(sheet_group_col, dropna=False):
            sheets.append((_safe_sheet_name(seed), g.copy()))

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="left", vertical="center")
    cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    used_names = set()

    for sheet_name, sdf in sheets:
        ws = wb.create_sheet(title=_safe_sheet_name(sheet_name))

        for c in out_cols:
            if c not in sdf.columns:
                sdf[c] = ""
        sdf = sdf[out_cols].copy()

        # header
        for col_idx, col_name in enumerate(out_cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # rows
        for r_idx, row in enumerate(sdf.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value="" if value is None else str(value))
                cell.alignment = cell_align

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

        # add Excel table
        if ws.max_row >= 2:
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            base_name = re.sub(r"[^A-Za-z0-9_]", "", f"Tbl_{ws.title}")[:22] or "Table"
            tname, i = base_name, 2
            while tname in used_names:
                tname = f"{base_name}{i}"
                i += 1
            used_names.add(tname)

            tab = Table(displayName=tname, ref=ref)
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False,
            )
            ws.add_table(tab)

        _autofit_columns(ws)

        if band_by_col in out_cols and ws.max_row >= 2:
            team_col_idx = out_cols.index(band_by_col) + 1
            _apply_team_banding(ws, team_col_idx, start_row=2, end_row=ws.max_row)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =============================================================================
# Streamlit UI
# =============================================================================

st.set_page_config(page_title="AR Directory Extractor", layout="wide")

st.markdown("""
<style>
.block-container { max-width: 1100px; padding-top: 2.2rem; padding-bottom: 3rem; }
html, body, [class*="css"]  { font-family: -apple-system, BlinkMacSystemFont, "SF Pro Text", "Segoe UI", Roboto, Helvetica, Arial, sans-serif; }
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
:root{
  --txt: rgba(0,0,0,0.92);
  --sub: rgba(0,0,0,0.62);
  --card-bg: rgba(255,255,255,0.72);
  --card-border: rgba(0,0,0,0.10);
  --shadow: rgba(0,0,0,0.08);
}
@media (prefers-color-scheme: dark){
  :root{
    --txt: rgba(255,255,255,0.92);
    --sub: rgba(255,255,255,0.68);
    --card-bg: rgba(25,25,25,0.55);
    --card-border: rgba(255,255,255,0.12);
    --shadow: rgba(0,0,0,0.35);
  }
}
.card {
  border: 1px solid var(--card-border);
  border-radius: 16px;
  padding: 18px 18px;
  background: var(--card-bg);
  box-shadow: 0 6px 24px var(--shadow);
  backdrop-filter: blur(8px);
}
.h1 { font-size: 40px; font-weight: 700; letter-spacing: -0.02em; margin: 0; color: var(--txt); }
.sub { font-size: 15px; margin-top: 6px; color: var(--sub); }
.stTextArea textarea, .stTextInput input { border-radius: 12px !important; }
.stButton button, .stDownloadButton button {
  border-radius: 999px !important;
  padding: 0.55rem 1.0rem !important;
  font-weight: 600 !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown('<p class="h1">AR Directory Extractor</p>', unsafe_allow_html=True)
st.markdown(
    '<p class="sub">Collects publicly visible advisor contact info. '
    '<b>TD directories:</b> pulls both <b>Advisors</b> + <b>Teams</b>, crawls profiles + rosters, then globally de-dupes people (no duplicate lines).</p>',
    unsafe_allow_html=True
)
st.write("")

left, right = st.columns([2.15, 1], gap="large")

with right:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("### Settings")
    sleep_s = st.slider("Polite delay (seconds)", 0.25, 2.0, 0.75, 0.25)
    max_targets = st.number_input("Max targets per run", 1, 600, 200, 10)
    drop_no_contact = st.checkbox("Drop rows with no email AND no phone", value=False)
    if not HAS_OPENPYXL:
        st.caption("Excel export disabled (openpyxl missing). Add it to requirements.txt.")
    st.markdown('</div>', unsafe_allow_html=True)

with left:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("### 1) Paste branch URLs")
    seed_urls_text = st.text_area(
        "Seed URLs (one per line)",
        height=140,
        placeholder="Example TD directory:\nhttps://advisors.td.com/montreal1/\n\nMultiple links supported (Excel creates 1 sheet per link)."
    )
    c1, c2 = st.columns([1, 1])
    discover_clicked = c1.button("Discover targets", type="primary", use_container_width=True)
    clear_clicked = c2.button("Clear", use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

if clear_clicked:
    for k in ["df_candidates", "edited_candidates", "df_clean", "errs_build"]:
        st.session_state.pop(k, None)
    st.rerun()

if discover_clicked:
    seeds = [s.strip() for s in seed_urls_text.splitlines() if s.strip()]
    if not seeds:
        st.warning("Paste at least one seed URL.")
    else:
        dfs, errors = [], []
        with st.spinner("Discovering targets..."):
            for s in seeds:
                try:
                    dfs.append(discover_targets_from_seed(s, sleep_s=sleep_s))
                except Exception as e:
                    errors.append({"seed": s, "error": str(e)})

        df_candidates = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame(
            columns=["branch_seed_url", "target_url", "link_text", "td_kind"]
        )
        if "td_kind" not in df_candidates.columns:
            df_candidates["td_kind"] = ""

        df_candidates = df_candidates.drop_duplicates(subset=["target_url", "td_kind"]).reset_index(drop=True)
        df_candidates["include"] = True

        st.session_state["df_candidates"] = df_candidates
        st.session_state.pop("edited_candidates", None)
        st.session_state.pop("df_clean", None)
        st.session_state["errs_build"] = errors

if "df_candidates" in st.session_state and not st.session_state["df_candidates"].empty:
    st.write("")
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("### 2) Select targets")
    dfc = st.session_state["df_candidates"].copy()

    a, b, c, _ = st.columns([1, 1, 1.2, 5])
    if a.button("Select all"):
        dfc["include"] = True
    if b.button("Select none"):
        dfc["include"] = False
    if c.button("Keep first 150"):
        dfc["include"] = False
        dfc.loc[:149, "include"] = True

    edited = st.data_editor(
        dfc[["include", "branch_seed_url", "target_url", "td_kind", "link_text"]],
        use_container_width=True,
        num_rows="dynamic"
    )
    st.session_state["edited_candidates"] = edited
    st.markdown('</div>', unsafe_allow_html=True)

build_ready = "edited_candidates" in st.session_state and not st.session_state["edited_candidates"].empty

st.write("")
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown("### 3) Build directory")

if not build_ready:
    st.info("Discover and select targets above, then build.")
else:
    edited = st.session_state["edited_candidates"].copy()
    chosen = edited[edited["include"] == True].head(int(max_targets)).copy()

    # TD: process Advisors first, then Teams
    order_map = {"advisor": 0, "team": 1, "": 2}
    chosen["_ord"] = chosen.get("td_kind", "").astype(str).str.lower().map(order_map).fillna(2).astype(int)
    chosen = chosen.sort_values(["_ord", "branch_seed_url", "target_url"]).drop(columns=["_ord"], errors="ignore")

    m1, m2, m3 = st.columns(3)
    m1.metric("Selected targets", int(len(chosen)))
    m2.metric("Polite delay", f"{sleep_s:.2f}s")
    m3.metric("Max this run", int(max_targets))

    build_clicked = st.button("Build directory", type="primary")

    if build_clicked:
        rows, errs = [], []
        prog = st.progress(0)
        total = len(chosen)

        with st.spinner("Building directory..."):
            for i, r in enumerate(chosen.itertuples(index=False), start=1):
                try:
                    td_kind = (getattr(r, "td_kind", "") or "").strip().lower()
                    target_url = getattr(r, "target_url", "")

                    # ---------------- TD ----------------
                    if is_td_url(target_url):
                        root_html, root_final = polite_get(target_url, sleep_s=sleep_s)

                        if not td_kind:
                            td_kind = td_detect_single_root_kind(root_html)

                        team_slug = to_team_slug(root_final)
                        team_name = page_title(root_html) or team_slug

                        # Advisors: crawl profile
                        if td_kind == "advisor":
                            people, src = td_fetch_people(root_final, sleep_s=sleep_s)
                            team_aff_name, team_aff_root = td_extract_part_of_team(root_html, root_final)
                            if team_aff_root:
                                team_name = team_aff_name or team_name

                            for p in people:
                                rows.append({
                                    "branch_seed_url": r.branch_seed_url,
                                    "team_root_url": team_aff_root or root_final,
                                    "team_slug": to_team_slug(team_aff_root or root_final),
                                    "team_name": team_name,
                                    "team_page_url": team_aff_root or root_final,
                                    "contact_page_url": "",
                                    "advisor_name": p.get("advisor_name", ""),
                                    "advisor_role": p.get("advisor_role", ""),
                                    "advisor_email": p.get("advisor_email", ""),
                                    "advisor_phone": p.get("advisor_phone", ""),
                                    "advisor_address": p.get("advisor_address", ""),
                                    "advisor_profile_url": p.get("advisor_profile_url", ""),
                                    "source": p.get("source", ""),
                                    "source_page_used": src,
                                })

                        # Teams: crawl roster
                        else:
                            meet_url = td_guess_meet_the_team_url(root_final, sleep_s=sleep_s) or root_final
                            people, src = td_fetch_people(meet_url, sleep_s=sleep_s)
                            if not people:
                                people, src = td_fetch_people(root_final, sleep_s=sleep_s)

                            for p in people:
                                rows.append({
                                    "branch_seed_url": r.branch_seed_url,
                                    "team_root_url": root_final,
                                    "team_slug": team_slug,
                                    "team_name": team_name,
                                    "team_page_url": meet_url,
                                    "contact_page_url": "",
                                    "advisor_name": p.get("advisor_name", ""),
                                    "advisor_role": p.get("advisor_role", ""),
                                    "advisor_email": p.get("advisor_email", ""),
                                    "advisor_phone": p.get("advisor_phone", ""),
                                    "advisor_address": p.get("advisor_address", ""),
                                    "advisor_profile_url": p.get("advisor_profile_url", ""),
                                    "source": p.get("source", ""),
                                    "source_page_used": src,
                                })

                        prog.progress(min(1.0, i / max(1, total)))
                        continue

                    # ---------------- Non-TD ----------------
                    html_root, root_final = polite_get(target_url, sleep_s=sleep_s)
                    slug = to_team_slug(root_final)
                    team_name = page_title(html_root) or slug

                    links = extract_links(html_root, root_final)
                    team_page = find_best_link(links, root_final, TEAM_PAGE_TEXT_PAT)
                    contact_page = find_best_link(links, root_final, CONTACT_PAGE_TEXT_PAT)

                    if is_desjardins_url(root_final):
                        team_page = root_final
                        contact_page = ""

                    if is_cibc_wg_url(root_final):
                        web_team_guess = f"https://woodgundyadvisors.cibc.com/web/{slug}/our-team"
                        web_contact_guess = f"https://woodgundyadvisors.cibc.com/web/{slug}/contact"
                        if not team_page:
                            try:
                                _, u = polite_get(web_team_guess, sleep_s=sleep_s)
                                team_page = u
                            except Exception:
                                pass
                        if not contact_page:
                            try:
                                _, u = polite_get(web_contact_guess, sleep_s=sleep_s)
                                contact_page = u
                            except Exception:
                                pass

                    people, source_page_used = [], ""

                    if team_page:
                        people, source_page_used = fetch_people(team_page, sleep_s=sleep_s)

                    if contact_page:
                        contact_people, contact_src = fetch_people(contact_page, sleep_s=sleep_s)
                        by_name = {canon_name(p.get("advisor_name", "")): p for p in people if p.get("advisor_name")}
                        for cp in contact_people:
                            k = canon_name(cp.get("advisor_name", ""))
                            if not k:
                                continue
                            if k in by_name:
                                for fld in ["advisor_email", "advisor_phone", "advisor_address", "advisor_role", "advisor_profile_url"]:
                                    if fld == "advisor_role":
                                        if (not by_name[k].get(fld)) and is_likely_role(cp.get(fld, ""), cp.get("advisor_name", "")):
                                            by_name[k][fld] = cp.get(fld, "")
                                    else:
                                        if not by_name[k].get(fld) and cp.get(fld):
                                            by_name[k][fld] = cp.get(fld)
                            else:
                                people.append(cp)

                    if not people:
                        people = extract_people_from_page(html_root, root_final)
                        source_page_used = root_final

                    for p in (people or [{}]):
                        rows.append({
                            "branch_seed_url": r.branch_seed_url,
                            "team_root_url": root_final,
                            "team_slug": slug,
                            "team_name": team_name,
                            "team_page_url": team_page,
                            "contact_page_url": contact_page,
                            "advisor_name": p.get("advisor_name", ""),
                            "advisor_role": p.get("advisor_role", ""),
                            "advisor_email": p.get("advisor_email", ""),
                            "advisor_phone": p.get("advisor_phone", ""),
                            "advisor_address": p.get("advisor_address", ""),
                            "advisor_profile_url": p.get("advisor_profile_url", ""),
                            "source": p.get("source", "no_people_found" if not people else p.get("source", "")),
                            "source_page_used": source_page_used or team_page or root_final,
                        })

                except Exception as e:
                    errs.append({"target_url": getattr(r, "target_url", ""), "error": str(e)})

                prog.progress(min(1.0, i / max(1, total)))

        df_out = pd.DataFrame(rows)
        df_clean = post_process_directory(df_out, drop_no_contact=drop_no_contact)

        st.session_state["df_clean"] = df_clean
        st.session_state["errs_build"] = errs
        st.success("Done. Scroll down to export.")

st.markdown('</div>', unsafe_allow_html=True)


# =============================================================================
# Export (remove branch_seed_url and team_slug from output)
# =============================================================================

if "df_clean" in st.session_state:
    st.write("")
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("### Results & Export")

    df_clean = st.session_state["df_clean"]
    errs = st.session_state.get("errs_build", [])

    OUT_COLS = ["team_name", "advisor_name", "advisor_role", "advisor_email", "advisor_phone"]

    df_export = df_clean.copy()
    for c in OUT_COLS:
        if c not in df_export.columns:
            df_export[c] = ""
    df_export = df_export[OUT_COLS].copy()

    a1, a2, a3 = st.columns(3)
    a1.metric("Rows exported", int(len(df_export)))
    a2.metric("Teams (unique)", int(df_export["team_name"].nunique()) if len(df_export) else 0)
    a3.metric("Errors", int(len(errs)))

    st.dataframe(df_export, use_container_width=True, height=420)

    # CSV (UTF-8 BOM so French accents render correctly in Excel)
    csv_bytes = df_export.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "Download CSV (UTF-8)",
        data=csv_bytes,
        file_name="directory_output.csv",
        mime="text/csv",
        use_container_width=True
    )

    # Excel: auto-fit + team banding + one sheet per input link
    if HAS_OPENPYXL:
        df_excel_full = df_clean.copy()
        if "branch_seed_url" not in df_excel_full.columns:
            df_excel_full["branch_seed_url"] = "Seed"

        try:
            xlsx_bytes = build_styled_xlsx(
                df_full=df_excel_full,
                sheet_group_col="branch_seed_url",
                out_cols=OUT_COLS,      # no branch_seed_url, no team_slug
                band_by_col="team_name"
            )
            st.download_button(
                "Download Excel (Styled .xlsx)",
                data=xlsx_bytes,
                file_name="directory_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.caption("Excel includes auto-fit widths, team color banding, and one worksheet per input link + an All sheet.")
        except Exception as e:
            st.warning(f"Excel export failed: {e}")
    else:
        st.info("Excel export is disabled (openpyxl missing). Add openpyxl to requirements.txt to enable .xlsx.")

    if errs:
        with st.expander("Show errors"):
            st.dataframe(pd.DataFrame(errs), use_container_width=True)

    st.markdown('</div>', unsafe_allow_html=True)
